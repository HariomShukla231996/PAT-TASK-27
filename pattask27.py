# excel_funtions.py file

from openpyxl import load_workbook


class Hari_Excel_Function:


   def __init__(self, file_name, sheet_name):
       self.file = file_name
       self.sheet = sheet_name


   # get the row count of my Excel file
   def row_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_row


   # get the column count of my Excel file
   def column_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_column


   # read the data from the Excel file
   def read_data(self, row_number, column_number):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.cell(row=row_number, column=column_number).value


   # write the data into the Excel file
   def write_data(self, row_number, column_number, data):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       sheet.cell(row=row_number, column=column_number).value = data
       workbook.save(self.file)

 # Locators.py file

lass Web_Locators:
   url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"
   dashboard_url = "https://opensource-demo.orangehrmlive.com/web/index.php/dashboard/index"
   username_locator = "username"
   password_locator = "password"
   submit_button = '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/form[1]/div[3]/button[1]'
   logout_button = '//*[@id="app"]/div[1]/div[1]/header/div[1]/div[2]/ul/li/span/p'

# main.py file

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from Locators.locators import Web_Locators
from excel_functions import Hari_Excel_Function
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.action_chains import ActionChains


excel_file = Hari_Excel_Function("test_data.xlsx", "Sheet1")
rows = excel_file.row_count()


driver = webdriver.Firefox(service=Service(GeckoDriverManager().install()))
driver.implicitly_wait(10)
driver.maximize_window()
driver.get(Web_Locators().url)


for row in range(2, rows+1):
   username = excel_file.read_data(row, 6)
   password = excel_file.read_data(row, 7)


   driver.find_element(by=By.NAME, value=Web_Locators().username_locator).send_keys(username)
   driver.find_element(by=By.NAME, value=Web_Locators().password_locator).send_keys(password)
   driver.find_element(by=By.XPATH, value=Web_Locators().submit_button).click()


   driver.implicitly_wait(10)


   if Web_Locators().dashboard_url in driver.current_url:
       print("SUCCESS : Login with Username {a}".format(a=username))
       excel_file.write_data(row, 8, "TEST PASSED")
       action = ActionChains(driver)
       logout_button = driver.find_element(by=By.XPATH, value=Web_Locators().logout_button)
       action.click(on_element=logout_button).perform()
       driver.find_element(by=By.LINK_TEXT, value="Logout").click()
   elif Web_Locators().url in driver.current_url:
       print("FAIL : Login failure with Username {a}".format(a=username))
       excel_file.write_data(row, 8, "TEST FAIL")
       driver.refresh()


driver.quit()

